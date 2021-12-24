import os
import json
import openpyxl
import warnings
warnings.filterwarnings('ignore')
URL=['www.biomedcentral.com','link.springer.com','www.nature.com','author-welcome.nature.com-41598']


### 
# 每次运行需要修改
folder_path='first_round_test'
iteration_time=5
fastly_telecom_path='first_round_test/fastly-telecom-onLoadTimeResult1221.json'
google_telecom_path='first_round_test/google-telecom-onLoadTimeResult1221.json'
tencent_telecom_path='first_round_test/tecent-telecom-onLoadTimeResult1221.json'
fastly_netherlands_path='first_round_test/fastly-netherlands-onLoadTimeResult1221.json'
google_netherlands_path='first_round_test/google-netherlands-onLoadTimeResult1221.json'
tencent_netherlands_path='first_round_test/tencent-netherlands-onLoadTimeResult1221.json'
###

report_path='Testreport.xlsx'
if(os.path.exists(report_path)):
    os.remove(report_path)
workbook = openpyxl.Workbook()


#------------------------Fastly_Telecom----------------------------------------#
sheet1 = workbook.create_sheet(index=0, title="fastly-telecom100")
for i in range(4):
    sheet1.cell(row=1,column=i+1).value=URL[i]


fastly_file=open(fastly_telecom_path,'r')
fastly_json=json.load(fastly_file)


for i in range(4):
    for m in fastly_json:    
        if(m.get('pageUrl')==URL[i]):
            print(URL[i])
            cycletime=m["pageLoadTimeForEveryCycle"]
            for j in range(iteration_time):
               print(cycletime[j])
               sheet1.cell(row=j+2,column=i+1).value=cycletime[j]
            
#-----------------------Google_Telecom-----------------------------------------#   
sheet2 = workbook.create_sheet(index=0, title="google-telecom100")
for i in range(4):
    sheet2.cell(row=1,column=i+1).value=URL[i]


google_file=open(google_telecom_path,'r')
google_json=json.load(google_file)


for i in range(4):
    for m in google_json:    
        if(m.get('pageUrl')==URL[i]):
            print(URL[i])
            cycletime=m["pageLoadTimeForEveryCycle"]
            for j in range(iteration_time):
               print(cycletime[j])
               sheet2.cell(row=j+2,column=i+1).value=cycletime[j]


#-----------------------Tencent_Telecom-----------------------------------------#   
sheet3 = workbook.create_sheet(index=0, title="tencent-telecom100")
for i in range(4):
    sheet3.cell(row=1,column=i+1).value=URL[i]


tencent_file=open(tencent_telecom_path,'r')
tencent_json=json.load(tencent_file)


for i in range(4):
    for m in tencent_json:    
        if(m.get('pageUrl')==URL[i]):
            print(URL[i])
            cycletime=m["pageLoadTimeForEveryCycle"]
            for j in range(iteration_time):
               print(cycletime[j])
               sheet3.cell(row=j+2,column=i+1).value=cycletime[j]

workbook.save("Testreport.xlsx")

#------------------------Fastly_Netherlands----------------------------------------#

sheet4 = workbook.create_sheet(index=0, title="fastly-netherlands100")
for i in range(4):
    sheet4.cell(row=1,column=i+1).value=URL[i]


fastly_file=open(fastly_netherlands_path,'r')
fastly_json=json.load(fastly_file)

for i in range(4):
    for m in fastly_json:    
        if(m.get('pageUrl')==URL[i]):
            print(URL[i])
            cycletime=m["pageLoadTimeForEveryCycle"]
            for j in range(iteration_time):
               print(cycletime[j])
               sheet4.cell(row=j+2,column=i+1).value=cycletime[j]

#-----------------------Google_Netherlands-----------------------------------------#   
sheet5 = workbook.create_sheet(index=0, title="google-netherlands100")
for i in range(4):
    sheet5.cell(row=1,column=i+1).value=URL[i]


google_file=open(google_netherlands_path,'r')
google_json=json.load(google_file)


for i in range(4):
    for m in google_json:    
        if(m.get('pageUrl')==URL[i]):
            print(URL[i])
            cycletime=m["pageLoadTimeForEveryCycle"]
            for j in range(iteration_time):
               print(cycletime[j])
               sheet5.cell(row=j+2,column=i+1).value=cycletime[j]

#-----------------------Tencent_Netherlands-----------------------------------------#   
sheet6 = workbook.create_sheet(index=0, title="tencent-netherlands100")
for i in range(4):
    sheet6.cell(row=1,column=i+1).value=URL[i]


tencent_file=open(tencent_netherlands_path,'r')
tencent_json=json.load(tencent_file)


for i in range(4):
    for m in tencent_json:    
        if(m.get('pageUrl')==URL[i]):
            print(URL[i])
            cycletime=m["pageLoadTimeForEveryCycle"]
            for j in range(iteration_time):
               print(cycletime[j])
               sheet6.cell(row=j+2,column=i+1).value=cycletime[j]

workbook.save("Testreport.xlsx")

